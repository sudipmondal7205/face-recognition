import cv2
from deepface import DeepFace
import os
import time
import openpyxl as xl

wb = xl.load_workbook(r'C:\Users\sudip\OneDrive\Programme\Projects\Python\FaceRecognition\Attendence.xlsx')
sheet = wb.active
row = sheet['A']
names = [cell.value for cell in row]
data_base = r'C:\Users\sudip\OneDrive\Programme\Projects\Python\FaceRecognition\images'
date = time.strftime(r'%d-%m-%Y')
images = os.listdir(r'C:\Users\sudip\OneDrive\Programme\Projects\Python\FaceRecognition\images')
for item in images:
    if item.endswith('.pkl') :
        images.remove(item)
for i in range(len(images)) :
    images[i] = images[i][:-4]
print(images)
for name in images:
    if not(names.count(name)) :
        names.append(name)
        sheet.append([name])
print(names)
Attendence_list = [date] + len(images)*['A']
cam = cv2.VideoCapture(0)
print("Opening webcam..... Press 'q' to quit....")

while True :
    camopened, frame = cam.read()

    if not(camopened) :
        print("Failed to capture frame.....  Exiting.....")
        break

    try :
        match = DeepFace.find(img_path = frame, db_path = data_base, enforce_detection = True)
        print('Match Found : ', match[0].iloc[0,0].split('\\')[-1][0:-4])
        if len(match[0]) > 0 :
            name = match[0].iloc[0,0].split('\\')[-1][0:-4]
            text = name
            color = (0, 255, 0)
            Time = time.strftime('%I-%M-%S')
            Attendence_list[names.index(name)] = Time
        else :
            text = 'NO MATCH'
            color = (0, 0, 255)
    except :
        text = 'NO FACE DETECTED'
        color = (0, 255, 255)

    cv2.putText(frame, text, (25, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2, cv2.LINE_AA)
    cv2.imshow('Live Face Recognition', frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

last_column = sheet.max_column
i=1
for item in Attendence_list:
    sheet.cell(row=i, column=last_column + 1, value=item)
    i += 1
wb.save(r'C:\Users\sudip\OneDrive\Programme\Projects\Python\FaceRecognition\Attendence.xlsx')
cam.release()
cv2.destroyAllWindows()